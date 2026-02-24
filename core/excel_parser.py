"""Liest eine LAF-Excel-Datei ein und gibt ein ExcelData-Objekt zurück."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from .models import ExcelData


class ExcelParseError(Exception):
    """Wird bei einem nicht behebbaren Fehler beim Excel-Parsing geworfen."""


_SPALTEN = {
    "gliederung": 0,
    "descriptor": 1,
    "label": 2,
    "alternatives_label": 3,
    "pflichtfeld": 4,
    "schreibgeschuetzt": 5,
}

_DATEN_AB_ZEILE = 4


def _lese_zellenwert(ws, zeile: int, spalte: int):
    wert = ws.cell(row=zeile, column=spalte + 1).value
    if wert == "" or wert is None:
        return None
    return wert


def parse_excel(pfad: str) -> ExcelData:
    """
    Liest eine LAF-formatierte Excel-Datei (.xlsx) ein.

    Raises:
        ExcelParseError: Bei fehlender Datei, korruptem Format oder fehlendem Inhalt.
    """
    excel_pfad = Path(pfad).resolve()

    if not excel_pfad.exists():
        raise ExcelParseError(f"Excel-Datei nicht gefunden: {excel_pfad}")

    if excel_pfad.suffix.lower() not in {".xlsx", ".xls"}:
        raise ExcelParseError(f"Ungültiges Dateiformat (erwartet .xlsx/.xls): {excel_pfad.name}")

    try:
        wb = openpyxl.load_workbook(str(excel_pfad), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f"Excel-Datei konnte nicht geöffnet werden: {exc}") from exc

    ws = wb.worksheets[0]
    blattname = ws.title

    formularname = ws.cell(row=1, column=2).value or ws.cell(row=1, column=1).value
    dateiname = ws.cell(row=3, column=2).value or ws.cell(row=3, column=1).value

    if not formularname:
        wb.close()
        raise ExcelParseError(f"Formularname fehlt in Zeile 1 von '{excel_pfad.name}'.")

    formularname = str(formularname).strip()
    dateiname = str(dateiname).strip() if dateiname else excel_pfad.stem

    zeilen: list[dict] = []
    for zeile_nr in range(_DATEN_AB_ZEILE, (ws.max_row or 0) + 1):
        zeile_dict = {
            spalte: _lese_zellenwert(ws, zeile_nr, idx)
            for spalte, idx in _SPALTEN.items()
        }
        if any(v is not None for v in zeile_dict.values()):
            zeilen.append(zeile_dict)

    wb.close()

    if not zeilen:
        raise ExcelParseError(f"Keine Datenzeilen gefunden in '{excel_pfad.name}'.")

    return ExcelData(
        formularname=formularname,
        dateiname=dateiname,
        blattname=blattname,
        zeilen=zeilen,
    )

